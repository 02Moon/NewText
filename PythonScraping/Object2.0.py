import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import ttk
import tkinter.font as tkFont
from datetime import date
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
import time
import tkinter.messagebox as messagebox


# 爬取数据函数
def get_data(page):
    data = []
    url = 'http://q.10jqka.com.cn//index/index/board/all/field/zdf/order/desc/page/{}'
    headers = {'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0'}
    for i in range(1,page + 1):
        html = requests.get(url=url.format(i), headers=headers)
        soup = BeautifulSoup(html.text, 'html.parser')
        table = soup.find('table', {'class': 'm-table m-pager-table'})
        trs = table.find_all('tr')
        for tr in trs:
            tds = tr.find_all('td')
            if(len(tds)>0):
                code = tds[1].text.strip()
                name = tds[2].text.strip()
                price = tds[3].text.strip()
                per = tds[4].text.strip()
                data.append([code, name, price, '{}%'.format(per)])
    return data

# 保存数据函数
def savefile(data):
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:D1')
    ws['A1'] = date.today()
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    for item in data:
        ws.append(item)
    wb.save('{} {}：{}.xlsx'.format(date.today(), datetime.now().hour, datetime.now().minute))
    messagebox.showinfo('提示', '保存成功')



# 主窗口函数
def main_form():
    window = tk.Tk()
    window.title('')
    # window.geometry('200x110')
    label = tk.Label(window, text='              请输入内容：              ')
    label.pack(pady=5)
    entry = tk.Entry(window)
    entry.pack(pady=5)
    frame1 = tk.Frame(window)
    frame1.pack()
    
    # 数据显示窗口函数
    def form(value):
        # window = tk.Tk()
        time_label = tk.Label(window, font=('Arial', 9))
        time_label.pack()
        
        # 窗口时间实时更新函数
        def update_time():
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            time_label.config(text=current_time)
            time_label.after(1000, update_time)

        update_time()
        table_t = ttk.Treeview(window, columns=("column1", "column2", "column3", "column4"), show="headings")
        table_t.pack()
        table_t.heading("column1", text="Code")
        table_t.heading("column2", text="名称")
        table_t.heading("column3", text="现价")
        table_t.heading("column4", text="涨跌幅")
        table_t.column("column1", minwidth=0, width=tkFont.Font().measure("1234567"), anchor='center')
        table_t.column("column2", minwidth=0, width=tkFont.Font().measure("文字文字文"))
        table_t.column("column3", minwidth=0, width=tkFont.Font().measure("0000.00"), anchor='center')
        table_t.column("column4", minwidth=0, width=tkFont.Font().measure("000.00%"), anchor='center')
        # print(value)
        for item in get_data(value):
            table_t.insert("", tk.END, values=(item))
        # Button
        frame = tk.Frame(window)
        frame.pack()
        
        # 控件函数
        def button_click():
            savefile(get_data(value))

        # 控件函数
        def button1_click():
            window.quit()

        button = tk.Button(frame, text='Save', width=5, height=1, command=button_click)
        button1 = tk.Button(frame, text='Exit', width=5, height=1, command=button1_click)
        button.pack(side='left', padx=5, pady=5)
        button1.pack(side='left', padx=5, pady=5)
        return window.mainloop()

    # 控件函数
    def button_click():
        if (entry.get() == ''):
            messagebox.showinfo('提示', '请输入内容')
        else:
            value = int(entry.get())
            label.destroy()
            entry.destroy()
            frame1.destroy()
            time.sleep(1)
            form(value)

    # 控件函数
    def button1_click():
        window.quit()

    button = tk.Button(frame1, text='Search', width=5, height=1, command=button_click)
    button1 = tk.Button(frame1, text='Exit', width=5, height=1, command=button1_click)
    button.pack(side='left', padx=5, pady=5)
    button1.pack(side='left', padx=5, pady=5)
    return window.mainloop()

if __name__ == '__main__':
    main_form()
