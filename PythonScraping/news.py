from datetime import date, datetime
import requests
from bs4 import BeautifulSoup
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import ttk
import tkinter.font as tkFont





#开头打招呼，输出今天提起

#今天日期
today = date.today()

now = datetime.now()
hour = now.hour

if hour < 6:
    print("早上好！",'[{}]'.format(today))
elif hour < 12:
    print("上午好！",'[{}]'.format(today))
elif hour < 14:
    print("中午好！",'[{}]'.format(today))
elif hour < 18:
    print("下午好！",'[{}]'.format(today))
else:
    print("晚上好！",'[{}]'.format(today))


# r_list = pattern.findall(html)
# soup = BeautifulSoup(html,'html.parser')
# table = soup.find('table',{'id':'tzrlTb'})

#---------表格----------
wb = Workbook()
ws = wb.active
ws.merge_cells('A1:C1')
mon = input('输入年份+月份：')
ws['A1'] = '{}投资日历'.format(mon)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.append(['日期','星期','事件','影响板块'])
#-----------------------

url1 = 'http://comment.10jqka.com.cn/tzrl/getTzrlData.php?callback=callback_dt&type=data&date={}'.format(mon)
headers = {'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0'}
html = requests.get(url=url1, headers=headers).text
js = json.dumps(html)
js1 = json.loads(js)
# 提取 JSON 对象
json_str = js1.replace('callback_dt(', '').replace(');', '')
json_data = json.loads(json_str)
# print(json_data)
leng = len(json_data['data'])
# 提取有效信息
for i in range(0,leng):
    date = json_data['data'][i]['date']
    week = json_data['data'][i]['week']
    if len(json_data['data'][i]['events']) > 1:
        for j in range(0,len(json_data['data'][i]['events'])):
            event = json_data['data'][i]['events'][j][0]
            # # --------------------------------------------------------
            # # 判断条件重写
            # if json_data['data'][i]['concept'] != [[]]:
            #     print(json_data['data'][i]['concept'][0])
            #     concept = json_data['data'][i]['concept'][j][0]['name']
            # else:
            #     concept = json_data['data'][i]['field'][j][0]['name']
            # --------------------------------------------------------
            if json_data['data'][i]['concept'] != [[]]:
                if json_data['data'][i]['concept'][j] == []:
                    concept = json_data['data'][i]['field'][j][0]['name']
                else:
                    concept = json_data['data'][i]['concept'][j][0]['name']
            else:
                concept = json_data['data'][i]['field'][j][0]['name']
            print(date, week, event, concept)
            ws.append([date, week, event, concept])

    else:
        event = json_data['data'][i]['events'][0][0]
        #--------------------------------------------------------
        # 判断条件重写
        # if 'concept' in json_data['data'][i]:
        if json_data['data'][i]['concept'] != [[]]:
            concept = json_data['data'][i]['concept'][0][0]['name']
        else:
            concept = json_data['data'][i]['field'][0][0]['name']
        # --------------------------------------------------------
        print(date, week, event, concept)
        ws.append([date, week, event, concept])
wb.save('{}投资日历.xlsx'.format(today))
